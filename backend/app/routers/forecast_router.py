from fastapi import APIRouter, Depends
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession
from typing import Annotated
from fastapi import UploadFile, File, HTTPException
import csv, io, os
from datetime import date

from app.database import get_db
from app.models.forecast import MonthlyForecast
from app.models.position import Position
from app.schemas.forecast import ForecastCreate, ForecastBatchCreate, ForecastResponse

router = APIRouter(prefix="/api/forecasts", tags=["Forecasts"])


@router.get("/", response_model=list[ForecastResponse])
async def list_forecasts(start_date: str, end_date: str, db: Annotated[AsyncSession, Depends(get_db)]):
    from datetime import date
    s = date.fromisoformat(start_date)
    e = date.fromisoformat(end_date)
    result = await db.execute(
        select(MonthlyForecast).where(MonthlyForecast.date >= s, MonthlyForecast.date <= e)
    )
    items = []
    for f in result.scalars().all():
        pos_result = await db.execute(select(Position).where(Position.id == f.position_id))
        pos = pos_result.scalar_one_or_none()
        items.append(ForecastResponse(
            id=f.id, position_id=f.position_id, date=f.date,
            daily_volume=f.daily_volume,
            position_name=pos.name if pos else None,
            created_at=f.created_at, updated_at=f.updated_at,
        ))
    return items


@router.post("/batch")
async def batch_create(data: ForecastBatchCreate, db: Annotated[AsyncSession, Depends(get_db)]):
    for item in data.items:
        result = await db.execute(
            select(MonthlyForecast).where(
                MonthlyForecast.position_id == item.position_id,
                MonthlyForecast.date == item.date,
            )
        )
        existing = result.scalar_one_or_none()
        if existing:
            existing.daily_volume = item.daily_volume
        else:
            db.add(MonthlyForecast(position_id=item.position_id, date=item.date, daily_volume=item.daily_volume))
    await db.flush()
    return {"ok": True}


@router.post("/import")
async def import_forecasts(db: Annotated[AsyncSession, Depends(get_db)], file: UploadFile = File(...)):
    if not file.filename:
        raise HTTPException(400, detail="No file provided")

    ext = os.path.splitext(file.filename)[1].lower()
    if ext not in (".csv", ".xlsx", ".xls"):
        raise HTTPException(400, detail="Only CSV and Excel (.xlsx/.xls) files are supported")

    content = await file.read()
    rows = []

    if ext == ".csv":
        text = content.decode("utf-8-sig")
        reader = csv.DictReader(io.StringIO(text))
        for row in reader:
            rows.append(row)
    else:
        try:
            import openpyxl
        except ImportError:
            raise HTTPException(500, detail="openpyxl is required for Excel import")

        wb = openpyxl.load_workbook(io.BytesIO(content), read_only=True)
        ws = wb.active
        header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        for row in ws.iter_rows(min_row=2, values_only=True):
            rows.append(dict(zip(header, row)))

    pos_result = await db.execute(select(Position))
    positions = {p.name: p.id for p in pos_result.scalars().all()}

    imported = 0
    for row in rows:
        date_str = str(row.get("date", "") or "").strip()
        pos_id_str = str(row.get("position_id", "") or "").strip()
        pos_name = str(row.get("position_name", "") or "").strip()
        vol_str = str(row.get("daily_volume", "") or "").strip()

        if not date_str or not vol_str:
            continue

        try:
            d = date.fromisoformat(date_str)
        except (ValueError, TypeError):
            continue

        try:
            vol = int(float(vol_str))
        except (ValueError, TypeError):
            continue

        if pos_id_str and pos_id_str.isdigit():
            pos_id = int(pos_id_str)
        elif pos_name and pos_name in positions:
            pos_id = positions[pos_name]
        else:
            continue

        exist = await db.execute(
            select(MonthlyForecast).where(
                MonthlyForecast.position_id == pos_id,
                MonthlyForecast.date == d,
            )
        )
        existing = exist.scalar_one_or_none()
        if existing:
            existing.daily_volume = vol
        else:
            db.add(MonthlyForecast(position_id=pos_id, date=d, daily_volume=vol))
        imported += 1

    await db.flush()
    return {"ok": True, "imported": imported, "filename": file.filename}
